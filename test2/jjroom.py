from urllib import request
from bs4 import BeautifulSoup
import json
import random
import os
from openpyxl import Workbook
from openpyxl import load_workbook

area_name = ['武昌', '洪山', '东湖高新', '江夏', '江岸', '东西湖', '汉阳']
area_name_code = {'武昌': 13, '洪山': 8, '东湖高新': 13,
                  '江夏': 15, '江岸': 12, '东西湖': 10, '汉阳': 10}
area_wh = {}

host = 'http://www.jiangroom.com/queryRooms.html'
rooms_url = 'http://www.jiangroom.com/queryRoomsAsync?'

charset = 'utf-8'
debug = True

excel_save_path = 'rooms.xlsx'


def main():
    if debug:
        for x in range(0, 1):
            getAllData(x)


def getAreaCode(offset):
    params = 'offset=%d' % offset
    with request.urlopen(rooms_url + params) as resp:
        data = resp.read().decode(charset)
        bean = json.loads(data)
    for i in bean:
        name = i.get('premiseAddress')
        for j in area_name:
            if(j in name):
                area_wh[j] = int(i.get('usableArea'))
                break
    if len(area_wh) == len(area_name):
        print(area_wh.items())
        return
    else:
        offset += 1
        print('操作了' + str(offset + 1) + "次,area_wh=" + str(len(area_wh)))
        getAreaCode(offset)


def getAllData(page):
    rooms = []
    params = 'offset=%d' % page
    req = request.Request(rooms_url, None, createHeaderIp())
    with request.urlopen(req) as resp:
        data = resp.read().decode(charset)
        bean = json.loads(data)
    for i in bean:
        rooms.append([i.get('bedroomNameAbbr'), i.get(
            'realityPrice'), i.get('usableArea')])

    save2Excel(rooms)


def createHeaderIp():
    headers = {}
    ip = '%s.%s.%s.%s' % (random.randint(0, 255), random.randint(0, 255),
                          random.randint(0, 255), random.randint(0, 255))
    headers['X-Forwarded-For'] = ip
    return headers


def save2Excel(rlists):
    ws0 = None
    ws1 = None
    wroom = None
    if os.path.exists(excel_save_path):
        wroom = load_workbook(excel_save_path)
    if wroom:
        ws0 = wroom['武汉租房']
        ws1 = wroom['武昌地区']
    else:
        wroom = Workbook()
        ws0 = wroom.create_sheet(title='武汉租房')
        ws1 = wroom.create_sheet(title='武昌地区')
        ws0.append(['序号', '小区', '价格'])
        ws1.append(['序号', '小区', '价格'])
    count = 1
    for i in range(len(rlists)):
        ws0.append([i + 1, rlists[i][0], rlists[i][1]])
        if rlists[i][2] == area_name_code['武昌']:
            ws1.append([count, rlists[i][0], rlists[i][1]])
            count += 1

    wroom.save(excel_save_path)

if __name__ == '__main__':
    main()
